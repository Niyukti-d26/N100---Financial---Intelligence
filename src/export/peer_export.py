# src/export/peer_export.py

from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.peer_engine.engine import PeerEngine


GREEN = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

YELLOW = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C"
)

RED = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)

GOLD = PatternFill(
    fill_type="solid",
    fgColor="FFD700"
)


class PeerExport:

    def __init__(self):

        self.engine = PeerEngine()

        self.df = self.engine.generate_peer_comparison()

        self.output_file = Path(
            "output/peer_comparison.xlsx"
        )

        self.output_file.parent.mkdir(
            parents=True,
            exist_ok=True
        )

    def export(self):

        workbook = Workbook()

        workbook.remove(
            workbook.active
        )
        
        print("Peer Groups Found:")
        print(
        sorted(
        self.df["peer_group_name"]
        .dropna()
        .unique()
    )
)
        peer_groups = sorted(
            self.df["peer_group_name"]
            .dropna()
            .unique()
        )

        percentile_columns = [

            "roe_percentile",
            "roce_percentile",
            "marketcap_percentile",
            "pe_percentile",
            "pb_percentile",
            "revenue_percentile",
            "pat_percentile",
            "eps_percentile",
            "asset_turnover_percentile",

            # Day 18 metrics
            "npm_percentile",
            "de_percentile",
            "fcf_percentile",
            "icr_percentile"
        ]

        export_columns = [

            "company_id",

            "return_on_equity_pct",
            "roe_rank",
            "roe_percentile",

            "return_on_capital_employed_pct",
            "roce_rank",
            "roce_percentile",

            "net_profit_margin_pct",
            "npm_rank",
            "npm_percentile",

            "debt_to_equity",
            "de_rank",
            "de_percentile",

            "free_cash_flow_cr",
            "fcf_rank",
            "fcf_percentile",

            "interest_coverage",
            "icr_rank",
            "icr_percentile",

            "market_cap_crore",
            "marketcap_rank",
            "marketcap_percentile",

            "pe_ratio",
            "pe_rank",
            "pe_percentile",

            "pb_ratio",
            "pb_rank",
            "pb_percentile",

            "revenue_cagr_5yr",
            "revenue_rank",
            "revenue_percentile",

            "pat_cagr_5yr",
            "pat_rank",
            "pat_percentile",

            "eps_cagr_5yr",
            "eps_rank",
            "eps_percentile",

            "asset_turnover",
            "asset_turnover_rank",
            "asset_turnover_percentile",

            "composite_quality_score",

            "is_benchmark"
        ]

        for group in peer_groups:

            sheet = workbook.create_sheet(
                title=group[:31]
            )

            group_df = self.df[
                self.df["peer_group_name"] == group
            ].copy()

            group_df = group_df.sort_values(
                "composite_quality_score",
                ascending=False
            )

            available_columns = [
                col
                for col in export_columns
                if col in group_df.columns
            ]

            group_df = group_df[
                available_columns
            ]

            # Header
            for col_num, col_name in enumerate(
                group_df.columns,
                start=1
            ):

                sheet.cell(
                    row=1,
                    column=col_num,
                    value=col_name
                )

            # Data
            for row_num, row in enumerate(
                group_df.values,
                start=2
            ):

                for col_num, value in enumerate(
                    row,
                    start=1
                ):

                    sheet.cell(
                        row=row_num,
                        column=col_num,
                        value=value
                    )

            # Percentile colours
            for percentile_col in percentile_columns:

                if percentile_col not in group_df.columns:
                    continue

                col_idx = (
                    group_df.columns.get_loc(
                        percentile_col
                    ) + 1
                )

                for row_idx in range(
                    2,
                    len(group_df) + 2
                ):

                    value = sheet.cell(
                        row=row_idx,
                        column=col_idx
                    ).value

                    if value is None:
                        continue

                    if value >= 75:

                        sheet.cell(
                            row=row_idx,
                            column=col_idx
                        ).fill = GREEN

                    elif value <= 25:

                        sheet.cell(
                            row=row_idx,
                            column=col_idx
                        ).fill = RED

                    else:

                        sheet.cell(
                            row=row_idx,
                            column=col_idx
                        ).fill = YELLOW

            # Benchmark row
            if "is_benchmark" in group_df.columns:

                benchmark_col = (
                    group_df.columns.get_loc(
                        "is_benchmark"
                    ) + 1
                )

                for row_idx in range(
                    2,
                    len(group_df) + 2
                ):

                    benchmark_value = sheet.cell(
                        row=row_idx,
                        column=benchmark_col
                    ).value

                    if benchmark_value == 1:

                        for col_idx in range(
                            1,
                            len(group_df.columns) + 1
                        ):

                            sheet.cell(
                                row=row_idx,
                                column=col_idx
                            ).fill = GOLD

            # Median row
            median_row = len(group_df) + 3

            sheet.cell(
                row=median_row,
                column=1,
                value="MEDIAN"
            )

            for col_idx, col_name in enumerate(
                group_df.columns,
                start=1
            ):

                if pd.api.types.is_numeric_dtype(
                    group_df[col_name]
                ):

                    sheet.cell(
                        row=median_row,
                        column=col_idx,
                        value=float(
                            group_df[col_name]
                            .median()
                        )
                    )

            # Auto column width
            for column in sheet.columns:

                max_length = 0

                column_letter = (
                    column[0].column_letter
                )

                for cell in column:

                    try:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                    except Exception:
                        pass

                sheet.column_dimensions[
                    column_letter
                ].width = max_length + 2

        workbook.save(
            self.output_file
        )

        print(
            f"Exported -> {self.output_file}"
        )
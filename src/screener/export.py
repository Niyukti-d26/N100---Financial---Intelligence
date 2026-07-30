from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

OUTPUT = Path("data/output")
OUTPUT.mkdir(parents=True, exist_ok=True)


def export_screeners(engine):
    """Function: export_screeners"""
    screeners = {
        "Quality Compounder": engine.quality_compounder(),
        "Value Pick": engine.value_pick(),
        "Growth Accelerator": engine.growth_accelerator(),
        "Dividend Champion": engine.dividend_champion(),
        "Debt Free Blue Chip": engine.debt_free_blue_chip(),
        "Turnaround Watch": engine.turnaround_watch(),
    }

    file_path = OUTPUT / "screener_output.xlsx"

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:

        for sheet, df in screeners.items():

            df = df.sort_values("composite_quality_score", ascending=False)

            df.to_excel(writer, sheet_name=sheet[:31], index=False)

    wb = load_workbook(file_path)

    green = PatternFill(fill_type="solid", start_color="C6EFCE")

    red = PatternFill(fill_type="solid", start_color="FFC7CE")

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        headers = [cell.value for cell in ws[1]]

        if "composite_quality_score" not in headers:
            continue

        score_col = headers.index("composite_quality_score") + 1

        for row in range(2, ws.max_row + 1):

            cell = ws.cell(row=row, column=score_col)

            if cell.value is None:
                continue

            if cell.value >= 70:
                cell.fill = green
            else:
                cell.fill = red

    wb.save(file_path)

    return file_path
